"""
Export utilities for CSV, Excel, and JSON formats
"""
import json
import csv
import io
from typing import Dict, Any, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def export_to_json(data: Dict[str, Any]) -> str:
    """Export data to JSON string"""
    return json.dumps(data, indent=2, ensure_ascii=False)

def export_to_csv(data: Dict[str, Any]) -> str:
    """Export data to CSV format (flat structure)"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['Field', 'Value', 'Confidence'])
    
    # Write extracted data
    extracted_data = data.get('extracted_data', {})
    field_confidence = data.get('field_confidence', {})
    
    for field, value in extracted_data.items():
        if field == 'line_items':
            # Handle line items separately
            for i, item in enumerate(value or []):
                for item_field, item_value in item.items():
                    writer.writerow([f'line_item_{i+1}_{item_field}', item_value, ''])
        else:
            confidence = field_confidence.get(field, 0) * 100
            writer.writerow([field, value, f'{confidence:.1f}%'])
    
    # Write summary info
    writer.writerow(['overall_confidence', f"{data.get('overall_confidence', 0) * 100:.1f}%", ''])
    writer.writerow(['math_validation', data.get('math_validation', {}).get('calculations_correct', False), ''])
    
    return output.getvalue()

def export_to_excel(data: Dict[str, Any]) -> bytes:
    """Export data to Excel format with formatting"""
    wb = Workbook()
    
    # Main data sheet
    ws_main = wb.active
    ws_main.title = "Invoice Data"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Headers
    headers = ['Field', 'Value', 'Confidence']
    for col, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    row = 2
    extracted_data = data.get('extracted_data', {})
    field_confidence = data.get('field_confidence', {})
    
    for field, value in extracted_data.items():
        if field == 'line_items':
            continue  # Handle in separate sheet
        
        confidence = field_confidence.get(field, 0) * 100
        ws_main.cell(row=row, column=1, value=field.replace('_', ' ').title())
        ws_main.cell(row=row, column=2, value=str(value))
        ws_main.cell(row=row, column=3, value=f'{confidence:.1f}%')
        row += 1
    
    # Summary
    ws_main.cell(row=row, column=1, value="Overall Confidence")
    ws_main.cell(row=row, column=2, value=f"{data.get('overall_confidence', 0) * 100:.1f}%")
    row += 1
    
    ws_main.cell(row=row, column=1, value="Math Validation")
    ws_main.cell(row=row, column=2, value=str(data.get('math_validation', {}).get('calculations_correct', False)))
    
    # Line items sheet
    line_items = extracted_data.get('line_items', [])
    if line_items:
        ws_items = wb.create_sheet("Line Items")
        
        # Headers for line items
        item_headers = ['Description', 'Quantity', 'Unit Price', 'Line Total']
        for col, header in enumerate(item_headers, 1):
            cell = ws_items.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Line item data
        for row, item in enumerate(line_items, 2):
            ws_items.cell(row=row, column=1, value=item.get('description', ''))
            ws_items.cell(row=row, column=2, value=item.get('quantity', ''))
            ws_items.cell(row=row, column=3, value=item.get('unit_price', ''))
            ws_items.cell(row=row, column=4, value=item.get('line_total', ''))
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def flatten_data_for_export(data: Dict[str, Any]) -> Dict[str, Any]:
    """Flatten nested data structure for easier export"""
    flattened = {}
    
    extracted_data = data.get('extracted_data', {})
    field_confidence = data.get('field_confidence', {})
    
    for field, value in extracted_data.items():
        if field == 'line_items':
            for i, item in enumerate(value or []):
                for item_field, item_value in item.items():
                    flattened[f'line_item_{i+1}_{item_field}'] = item_value
        else:
            flattened[field] = value
            flattened[f'{field}_confidence'] = field_confidence.get(field, 0) * 100
    
    # Add summary fields
    flattened['overall_confidence'] = data.get('overall_confidence', 0) * 100
    flattened['math_validation_passed'] = data.get('math_validation', {}).get('calculations_correct', False)
    flattened['missing_required_fields'] = ', '.join(data.get('missing_required_fields', []))
    
    return flattened